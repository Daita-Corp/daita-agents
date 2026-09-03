from __future__ import annotations

import asyncio
import sqlite3
import threading
import time as time_module
import warnings
from collections import defaultdict
from collections.abc import Mapping
from datetime import UTC, date, datetime, time, timedelta
from decimal import Decimal
from hashlib import sha256
from io import BytesIO
from pathlib import Path
from typing import Any
from uuid import UUID
from zipfile import ZIP_DEFLATED, ZipFile, ZipInfo

import openpyxl
import pytest
from _toolbox_model_support import (
    ToolboxAwareMockModelProvider as MockModelProvider,
)
from _workspace_support import workspace_for

from daita import Agent, SQLiteSource
from daita._json import canonical_json
from daita.adapters import (
    postgresql_query as postgresql_query_module,
    sqlite_query as sqlite_query_module,
)
from daita.artifacts.models import ArtifactAuthorship, ArtifactError
from daita.artifacts.renderers import (
    MAX_XLSX_BYTES,
    MAX_XLSX_COLUMNS,
    MAX_XLSX_MEMBERS,
    MAX_XLSX_ROWS,
    MAX_XLSX_SECONDS,
    MAX_XLSX_UNCOMPRESSED_BYTES,
    XLSX_MEDIA_TYPE,
    ExactXlsxProvenance,
    ExactXlsxRenderer,
    read_exact_xlsx_data,
    render_exact_xlsx,
    verify_exact_xlsx,
)
from daita.catalog.models import Sensitivity
from daita.domains.data.export_capabilities import (
    DATA_EXPORT_TABULAR_CAPABILITY_ID,
    DATA_EXPORT_TABULAR_TOOL_NAME,
    data_export_tabular_capability_declarations,
)
from daita.llm.models import (
    FinishReason,
    MessageRole,
    ModelProfile,
    ModelResponse,
    TextBlock,
    ToolCall,
    ToolResultBlock,
)
from daita.loop.models import LoopExitKind

_CREATED_AT = datetime(2026, 8, 1, 12, 30, 45, tzinfo=UTC)


def _ids():
    counts: defaultdict[str, int] = defaultdict(int)

    def create(prefix: str) -> str:
        counts[prefix] += 1
        if prefix in {"run", "conversation", "artifact", "destination"}:
            return f"{prefix}-{counts[prefix]:032x}"
        return f"{prefix}-{counts[prefix]}"

    return create


def _profile(provider: MockModelProvider) -> ModelProfile:
    return ModelProfile(
        id=provider.provider_id,
        context_window_tokens=32_000,
        max_output_tokens=2_000,
        supports_tools=True,
        supports_parallel_tools=True,
    )


def _tools(*calls: ToolCall) -> ModelResponse:
    return ModelResponse(finish_reason=FinishReason.TOOL_CALLS, tool_calls=calls)


def _stop(text: str = "done") -> ModelResponse:
    return ModelResponse(finish_reason=FinishReason.STOP, text=text)


def _error_code(block: ToolResultBlock) -> str:
    error = block.output.get("error")
    assert isinstance(error, Mapping)
    code = error.get("code")
    assert isinstance(code, str)
    return code


def _result_for_call(transcript, call_id: str) -> ToolResultBlock:
    return next(
        block
        for message in transcript.messages
        for block in message.content
        if isinstance(block, ToolResultBlock) and block.call_id == call_id
    )


def _provenance() -> ExactXlsxProvenance:
    return ExactXlsxProvenance(
        source_id="source-1",
        source_revision="schema_version:7",
        resource_revisions=(("resource-1", "sha256:" + "1" * 64),),
        sql_fingerprint="sha256:" + "2" * 64,
        parameters_sha256="sha256:" + "3" * 64,
        sensitivity=Sensitivity.CONFIDENTIAL,
        created_at=_CREATED_AT,
    )


def test_exact_xlsx_frozen_scalars_precision_and_fixed_provenance() -> None:
    content = render_exact_xlsx(
        (
            "null",
            "empty",
            "boolean",
            "integer",
            "large_integer",
            "float",
            "decimal",
            "date",
            "time",
            "timestamp",
            "binary",
            "uuid",
        ),
        (
            (
                None,
                "",
                True,
                999_999_999_999_999,
                1_000_000_000_000_000,
                -0.0,
                Decimal("12.3400"),
                date(2026, 8, 1),
                time(1, 2, 3, 4, tzinfo=UTC),
                datetime(2026, 8, 1, 1, 2, 3, 4, tzinfo=UTC),
                b"\x00\xff",
                UUID("ABCDEFAB-CDEF-ABCD-EFAB-CDEFABCDEFAB"),
            ),
        ),
        provenance=_provenance(),
    )

    verify_exact_xlsx(content)
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=False)
    assert workbook.sheetnames == ["Data", "Provenance"]
    data = workbook["Data"]
    assert data.cell(2, 1).value is None
    assert data.cell(2, 2).value == ""
    assert data.cell(2, 3).value is True
    assert data.cell(2, 4).value == 999_999_999_999_999
    assert data.cell(2, 4).data_type == "n"
    assert data.cell(2, 5).value == "1000000000000000"
    assert data.cell(2, 5).data_type == "s"
    assert data.cell(2, 6).value == 0
    assert data.cell(2, 7).value == "12.3400"
    assert data.cell(2, 8).value == datetime(2026, 8, 1)
    assert data.cell(2, 8).number_format == "yyyy-mm-dd"
    assert data.cell(2, 9).value == "01:02:03.000004+00:00"
    assert data.cell(2, 10).value == "2026-08-01T01:02:03.000004+00:00"
    assert data.cell(2, 11).value == r"\BAP8="
    assert data.cell(2, 12).value == "abcdefab-cdef-abcd-efab-cdefabcdefab"

    provenance = tuple(
        (
            workbook["Provenance"].cell(row, 1).value,
            workbook["Provenance"].cell(row, 2).value,
        )
        for row in range(1, workbook["Provenance"].max_row + 1)
    )
    assert provenance == (
        ("Key", "Value"),
        ("Authorship", "exact_source_data"),
        ("Source ID", "source-1"),
        ("Source Revision", "schema_version:7"),
        (
            "Resource Revisions",
            '[["resource-1","sha256:1111111111111111111111111111111111111111111111111111111111111111"]]',
        ),
        ("SQL Fingerprint", "sha256:" + "2" * 64),
        ("Parameters SHA-256", "sha256:" + "3" * 64),
        (
            "Columns SHA-256",
            "sha256:898e374845282024003ed783aaf2ee16ea1d9ea564e5b5584260c9e366b04675",
        ),
        ("Column Count", 12),
        ("Row Count", 1),
        ("Sensitivity", "confidential"),
        ("Created At", "2026-08-01T12:30:45Z"),
    )


def test_fixed_xlsx_reader_preserves_data_rows_including_all_blank_rows() -> None:
    content = render_exact_xlsx(
        ("label", "on_date", "enabled"),
        (
            ("alpha", date(2026, 8, 1), True),
            (None, None, None),
            ("omega", None, False),
        ),
        provenance=_provenance(),
    )

    data = read_exact_xlsx_data(content)

    assert data.columns == ("label", "on_date", "enabled")
    assert data.rows == (
        ("alpha", date(2026, 8, 1), True),
        (None, None, None),
        ("omega", None, False),
    )


def test_exact_xlsx_formula_and_url_like_text_remains_literal() -> None:
    content = render_exact_xlsx(
        ("formula", "url", "email"),
        (("=1+1", "https://example.test/path", "mailto:test@example.test"),),
        provenance=_provenance(),
    )
    verify_exact_xlsx(content)
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=False)
    row = workbook["Data"][2]
    assert tuple(cell.value for cell in row) == (
        "=1+1",
        "https://example.test/path",
        "mailto:test@example.test",
    )
    assert all(cell.data_type == "s" for cell in row)


def test_exact_xlsx_is_byte_deterministic_for_identical_input_and_time() -> None:
    first = render_exact_xlsx(
        ("name", "amount"),
        (("alpha", Decimal("1.20")),),
        provenance=_provenance(),
    )
    second = render_exact_xlsx(
        ("name", "amount"),
        (("alpha", Decimal("1.20")),),
        provenance=_provenance(),
    )
    assert first == second


@pytest.mark.parametrize(
    "value",
    (
        float("nan"),
        float("inf"),
        Decimal("NaN"),
        Decimal("Infinity"),
        [],
        {},
        (1, 2),
        timedelta(days=1),
        "\ud800",
        "😀" * 16_384,
        object(),
    ),
)
def test_exact_xlsx_rejects_unsupported_and_overlong_values_without_leakage(
    value: object,
) -> None:
    with pytest.raises(ArtifactError) as failure:
        render_exact_xlsx(("value",), ((value,),), provenance=_provenance())
    assert failure.value.code == "artifact_unsupported_value"
    assert failure.value.details["row_index"] == 0
    assert failure.value.details["column_index"] == 0
    assert failure.value.details["column_name"] == "value"
    assert "value" not in failure.value.details


def test_exact_xlsx_accepts_exact_utf16_text_limit_without_truncation() -> None:
    value = "😀" * 16_383 + "a"
    assert len(value.encode("utf-16-le")) // 2 == 32_767
    content = render_exact_xlsx(("value",), ((value,),), provenance=_provenance())
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=False)
    assert workbook["Data"]["A2"].value == value


def test_exact_xlsx_integer_precision_boundaries_are_frozen() -> None:
    values = (
        -999_999_999_999_999,
        999_999_999_999_999,
        -1_000_000_000_000_000,
        1_000_000_000_000_000,
        9_007_199_254_740_991,
        9_007_199_254_740_992,
    )
    content = render_exact_xlsx(
        tuple(f"c{index}" for index in range(len(values))),
        (values,),
        provenance=_provenance(),
    )
    row = openpyxl.load_workbook(BytesIO(content), data_only=False)["Data"][2]
    assert tuple(cell.data_type for cell in row) == ("n", "n", "s", "s", "s", "s")
    assert tuple(cell.value for cell in row) == (
        -999_999_999_999_999,
        999_999_999_999_999,
        "-1000000000000000",
        "1000000000000000",
        "9007199254740991",
        "9007199254740992",
    )


def test_exact_xlsx_row_column_byte_member_uncompressed_and_time_limits_fail_closed() -> (
    None
):
    assert (
        MAX_XLSX_ROWS,
        MAX_XLSX_COLUMNS,
        MAX_XLSX_BYTES,
        MAX_XLSX_UNCOMPRESSED_BYTES,
        MAX_XLSX_MEMBERS,
        MAX_XLSX_SECONDS,
    ) == (100_000, 256, 64 * 1024 * 1024, 256 * 1024 * 1024, 64, 60.0)

    with pytest.raises(ArtifactError) as row_failure:
        render_exact_xlsx(("x",), ((1,), (2,)), provenance=_provenance(), max_rows=1)
    assert row_failure.value.details["reason"] == "row_limit"

    with pytest.raises(ArtifactError) as column_failure:
        render_exact_xlsx(("a", "b"), (), provenance=_provenance(), max_columns=1)
    assert column_failure.value.code == "artifact_quota_exceeded"

    with pytest.raises(ArtifactError) as byte_failure:
        render_exact_xlsx(("x",), (), provenance=_provenance(), max_bytes=100)
    assert byte_failure.value.details["reason"] == "byte_limit"

    with pytest.raises(ArtifactError) as member_failure:
        render_exact_xlsx(("x",), (), provenance=_provenance(), max_members=10)
    assert member_failure.value.details["reason"] == "member_limit"

    with pytest.raises(ArtifactError) as uncompressed_failure:
        render_exact_xlsx(
            ("x",), (), provenance=_provenance(), max_uncompressed_bytes=100
        )
    assert uncompressed_failure.value.details["reason"] == "uncompressed_byte_limit"

    current = [0.0]
    renderer = ExactXlsxRenderer(
        ("x",), provenance=_provenance(), max_seconds=1, clock=lambda: current[0]
    )
    renderer.append((1,))
    current[0] = 2.0
    with pytest.raises(ArtifactError) as time_failure:
        renderer.finish()
    assert time_failure.value.details["reason"] == "time_limit"


def _rewrite_xlsx(
    content: bytes,
    *,
    replacements: Mapping[str, bytes] | None = None,
    duplicate: str | None = None,
) -> bytes:
    target = BytesIO()
    replacements = {} if replacements is None else replacements
    with (
        ZipFile(BytesIO(content), "r") as source,
        ZipFile(target, "w", compression=ZIP_DEFLATED) as destination,
    ):
        for source_info in source.infolist():
            raw = replacements.get(source_info.filename, source.read(source_info))
            info = ZipInfo(source_info.filename, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o600 << 16
            destination.writestr(info, raw)
            if source_info.filename == duplicate:
                with warnings.catch_warnings():
                    warnings.simplefilter("ignore", UserWarning)
                    destination.writestr(info, raw)
    return target.getvalue()


def test_exact_xlsx_production_verifier_rejects_formula_external_and_duplicate_zip() -> (
    None
):
    content = render_exact_xlsx(("value",), (("safe",),), provenance=_provenance())
    with ZipFile(BytesIO(content), "r") as archive:
        sheet = archive.read("xl/worksheets/sheet1.xml")
        relationships = archive.read("_rels/.rels")

    formula = sheet.replace(b"<v>1</v>", b"<f>1+1</f><v>1</v>", 1)
    with pytest.raises(ArtifactError) as formula_failure:
        verify_exact_xlsx(
            _rewrite_xlsx(
                content,
                replacements={"xl/worksheets/sheet1.xml": formula},
            )
        )
    assert formula_failure.value.details["reason"] == "unsafe_worksheet"

    external = relationships.replace(
        b"</Relationships>",
        (
            b'<Relationship Id="external" '
            b'Type="http://schemas.openxmlformats.org/officeDocument/2006/'
            b'relationships/hyperlink" Target="https://example.test" '
            b'TargetMode="External"/></Relationships>'
        ),
    )
    with pytest.raises(ArtifactError) as external_failure:
        verify_exact_xlsx(
            _rewrite_xlsx(content, replacements={"_rels/.rels": external})
        )
    assert external_failure.value.details["reason"] == "external_relationship"

    with pytest.raises(ArtifactError) as duplicate_failure:
        verify_exact_xlsx(_rewrite_xlsx(content, duplicate="xl/worksheets/sheet1.xml"))
    assert duplicate_failure.value.details["reason"] == "duplicate_member"


def test_exact_xlsx_source_failure_never_returns_a_partial_workbook() -> None:
    def rows():
        yield (1,)
        raise RuntimeError("source failed")

    with pytest.raises(RuntimeError, match="source failed"):
        render_exact_xlsx(("x",), rows(), provenance=_provenance())


def test_tabular_tool_schemas_add_xlsx_without_accepting_rows_bytes_or_workbook_control() -> (
    None
):
    declarations = data_export_tabular_capability_declarations()
    views = {
        item.name: item
        for item in declarations.tool_views
        if item.name == DATA_EXPORT_TABULAR_TOOL_NAME
    }
    assert set(views) == {DATA_EXPORT_TABULAR_TOOL_NAME}
    for view in views.values():
        capability = next(
            item for item in declarations.capabilities if item.id == view.capability_id
        )
        properties = capability.input_schema["properties"]
        assert isinstance(properties, Mapping)
        assert set(properties) == {
            "source_id",
            "resource_ids",
            "sql",
            "parameters",
            "format",
            "filename",
        }
        format_schema = properties["format"]
        assert isinstance(format_schema, Mapping)
        assert format_schema["type"] == "string"
        assert format_schema["enum"] == ("csv", "xlsx")
        assert set(properties).isdisjoint(
            {
                "rows",
                "content",
                "bytes",
                "provenance",
                "sensitivity",
                "sheet_name",
                "properties",
                "created_at",
            }
        )
        assert capability.artifact_policy is not None
        assert capability.artifact_policy.allowed_media_types == frozenset(
            {"text/csv", XLSX_MEDIA_TYPE}
        )


async def _sqlite_export_agent(
    tmp_path: Path,
    *,
    name: str,
    downloads: Path,
    rows: tuple[tuple[str, int], ...] = (("row", 1),),
) -> tuple[Agent, MockModelProvider, str, str, Path]:
    database = tmp_path / f"{name}.db"
    with sqlite3.connect(database) as connection:
        connection.execute("CREATE TABLE records(label TEXT, number INTEGER)")
        connection.executemany("INSERT INTO records VALUES (?, ?)", rows)
    provider = MockModelProvider((), provider_id=f"mock:{name}")
    agent = await Agent.create(
        name,
        root=tmp_path,
        model=provider,
        model_profile=_profile(provider),
        id_factory=_ids(),
        downloads_directory=downloads,
        clock=lambda: _CREATED_AT,
        workspace=workspace_for(tmp_path),
    )
    source = await agent.attach(SQLiteSource(database))
    resource = (await agent.list_catalog_resources(source_id=source.id))[0]
    return agent, provider, source.id, resource.id, database


async def test_sqlite_public_xlsx_creation_delivery_restart_and_redelivery(
    tmp_path: Path,
) -> None:
    downloads = tmp_path / "downloads-public-xlsx"
    downloads.mkdir()
    secret = "XLSX_ROW_SECRET_72a3d3"
    agent, provider, source_id, resource_id, _database = await _sqlite_export_agent(
        tmp_path,
        name="xlsx-public",
        downloads=downloads,
        rows=((secret, 1), ("=formula", 2)),
    )
    provider.replace_script(
        (
            _tools(
                ToolCall(
                    id="export",
                    name=DATA_EXPORT_TABULAR_TOOL_NAME,
                    arguments={
                        "source_id": source_id,
                        "resource_ids": (resource_id,),
                        "sql": "SELECT label, number FROM records ORDER BY number",
                        "format": "xlsx",
                        "filename": "records.xlsx",
                    },
                )
            ),
            _tools(
                ToolCall(
                    id="save",
                    name="artifact_save_local",
                    arguments={
                        "artifact_id": "artifact-00000000000000000000000000000001",
                        "mode": "create_new",
                        "destination_id": "default",
                    },
                )
            ),
            _stop("saved"),
        )
    )
    try:
        result = await agent.run("Export every record as an exact XLSX workbook.")
        assert result.kind is LoopExitKind.COMPLETED
        assert len(result.artifacts) == len(result.artifact_deliveries) == 1
        ref = result.artifacts[0]
        assert ref.call_id == "export"
        assert ref.capability_id == DATA_EXPORT_TABULAR_CAPABILITY_ID
        assert ref.media_type == XLSX_MEDIA_TYPE
        assert ref.filename == "records.xlsx"
        assert ref.sensitivity is Sensitivity.INTERNAL
        assert ref.provenance.authorship is ArtifactAuthorship.EXACT_SOURCE_DATA
        assert ref.provenance.columns == ("label", "number")
        assert ref.provenance.row_count == 2
        assert ref.provenance.parameters_sha256 == (
            "sha256:" + sha256(canonical_json(()).encode()).hexdigest()
        )

        payload = await agent.read_artifact(ref.artifact_id)
        verify_exact_xlsx(payload.content)
        workbook = openpyxl.load_workbook(BytesIO(payload.content), data_only=False)
        assert workbook.sheetnames == ["Data", "Provenance"]
        assert tuple(cell.value for cell in workbook["Data"][2]) == (secret, 1)
        assert tuple(cell.value for cell in workbook["Data"][3]) == ("=formula", 2)
        assert workbook["Data"]["A3"].data_type == "s"
        provenance = {
            workbook["Provenance"]
            .cell(row, 1)
            .value: workbook["Provenance"]
            .cell(row, 2)
            .value
            for row in range(2, workbook["Provenance"].max_row + 1)
        }
        assert provenance["Sensitivity"] == "internal"
        assert provenance["Row Count"] == 2
        assert provenance["Created At"] == "2026-08-01T12:30:45Z"
        receipt = result.artifact_deliveries[0]
        assert Path(receipt.saved_path).read_bytes() == payload.content

        transcript = await agent.transcript(result.run_id)
        serialized = canonical_json(
            [
                block.output
                for message in transcript.messages
                for block in message.content
                if isinstance(block, ToolResultBlock)
            ]
        )
        assert secret not in serialized
        assert secret not in repr(result)
        assert secret.encode() not in (agent.home / "state.db").read_bytes()
        first_request = "\n".join(
            block.text
            for message in provider.requests[0].messages
            for block in message.content
            if isinstance(block, TextBlock)
        )
        assert "data_export_tabular for exact CSV/XLSX" in first_request
        assert "Never put source rows or artifact bytes in arguments" in first_request
    finally:
        await agent.close()

    reopened = await Agent.open(
        "xlsx-public",
        root=tmp_path,
        downloads_directory=downloads,
        clock=lambda: _CREATED_AT,
        workspace=workspace_for(tmp_path),
    )
    try:
        ref = result.artifacts[0]
        payload = await reopened.read_artifact(ref.artifact_id)
        verify_exact_xlsx(payload.content)
        redelivery = await reopened.save_artifact(ref.artifact_id)
        assert Path(redelivery.saved_path).read_bytes() == payload.content
        assert redelivery.sha256 == ref.sha256
    finally:
        await reopened.close()


class _Attribute:
    def __init__(self, name: str) -> None:
        self.name = name


class _Cursor:
    def __init__(self, rows: tuple[tuple[object, ...], ...]) -> None:
        self._rows = rows
        self._index = 0

    def __aiter__(self):
        return self

    async def __anext__(self):
        if self._index >= len(self._rows):
            raise StopAsyncIteration
        value = self._rows[self._index]
        self._index += 1
        return value


class _Statement:
    def __init__(
        self,
        *,
        names: tuple[str, ...] = (),
        rows: tuple[tuple[object, ...], ...] = (),
    ) -> None:
        self._names = names
        self._rows = rows
        self.cursor_arguments: tuple[object, ...] | None = None

    def get_attributes(self):
        return tuple(_Attribute(name) for name in self._names)

    def cursor(self, *parameters: object, **kwargs: object):
        del kwargs
        self.cursor_arguments = parameters
        return _Cursor(self._rows)


class _Connection:
    def __init__(
        self,
        names: tuple[str, ...],
        rows: tuple[tuple[object, ...], ...],
    ) -> None:
        self.shape = _Statement(names=names)
        self.result = _Statement(rows=rows)
        self.prepared_sql: list[str] = []

    async def prepare(self, sql: str, *, timeout: float):
        del timeout
        self.prepared_sql.append(sql)
        return self.shape if len(self.prepared_sql) == 1 else self.result


async def test_postgresql_xlsx_adapter_streams_typed_values_without_json_or_csv() -> (
    None
):
    connection = _Connection(
        ("amount", "when", "payload"),
        (
            (
                Decimal("1.2300"),
                datetime(2026, 8, 1, tzinfo=UTC),
                b"abc",
                True,
            ),
        ),
    )
    content, columns, row_count = (
        await postgresql_query_module._execute_exact_tabular_query(
            connection,
            'SELECT amount, "when", payload FROM public.orders WHERE id = $1',
            (7,),
            format_name="xlsx",
            xlsx_provenance=_provenance(),
            max_rows=100_000,
            max_columns=256,
            max_bytes=64 * 1024 * 1024,
            timeout_seconds=60,
        )
    )
    verify_exact_xlsx(content)
    assert columns == ("amount", "when", "payload")
    assert row_count == 1
    workbook = openpyxl.load_workbook(BytesIO(content), data_only=False)
    assert tuple(cell.value for cell in workbook["Data"][2]) == (
        "1.2300",
        "2026-08-01T00:00:00.000000+00:00",
        r"\BYWJj",
    )
    assert connection.result.cursor_arguments == (7,)
    executed = connection.prepared_sql[1]
    assert "pg_catalog.pg_column_size" in executed
    assert "LIMIT 100001" in executed
    assert "to_json" not in executed.casefold()
    assert "jsonb" not in executed.casefold()
    assert "csv" not in executed.casefold()


async def test_concurrent_csv_xlsx_exports_keep_order_and_failed_siblings(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    downloads = tmp_path / "downloads-mixed-order"
    downloads.mkdir()
    agent, provider, source_id, resource_id, _database = await _sqlite_export_agent(
        tmp_path,
        name="mixed-order",
        downloads=downloads,
        rows=(("one", 1), ("two", 2)),
    )
    original = sqlite_query_module._run_exact_tabular_sync

    def delayed(*args: Any, **kwargs: Any):
        if "slow_export" in args[1]:
            time_module.sleep(0.1)
        return original(*args, **kwargs)

    monkeypatch.setattr(sqlite_query_module, "_run_exact_tabular_sync", delayed)
    provider.replace_script(
        (
            _tools(
                ToolCall(
                    id="first-xlsx",
                    name=DATA_EXPORT_TABULAR_TOOL_NAME,
                    arguments={
                        "source_id": source_id,
                        "resource_ids": (resource_id,),
                        "sql": "SELECT label FROM records /* slow_export */ ORDER BY number",
                        "format": "xlsx",
                        "filename": "first.xlsx",
                    },
                ),
                ToolCall(
                    id="failed",
                    name=DATA_EXPORT_TABULAR_TOOL_NAME,
                    arguments={
                        "source_id": source_id,
                        "resource_ids": (resource_id,),
                        "sql": "SELECT label AS duplicate, number AS duplicate FROM records",
                        "format": "xlsx",
                        "filename": "failed.xlsx",
                    },
                ),
                ToolCall(
                    id="third-csv",
                    name=DATA_EXPORT_TABULAR_TOOL_NAME,
                    arguments={
                        "source_id": source_id,
                        "resource_ids": (resource_id,),
                        "sql": "SELECT number FROM records ORDER BY number",
                        "format": "csv",
                        "filename": "third.csv",
                    },
                ),
            ),
            _stop(),
        )
    )
    try:
        result = await agent.run("Export these CSV and XLSX files.")
        transcript = await agent.transcript(result.run_id)
        blocks = tuple(
            message.content[0]
            for message in transcript.messages
            if message.role is MessageRole.TOOL
            and isinstance(message.content[0], ToolResultBlock)
            and message.content[0].output.get("kind") != "toolbox_load_receipt"
        )
        assert tuple(block.call_id for block in blocks) == (
            "first-xlsx",
            "failed",
            "third-csv",
        )
        assert _error_code(blocks[1]) == "artifact_unsupported_value"
        assert tuple(ref.call_id for ref in result.artifacts) == (
            "first-xlsx",
            "third-csv",
        )
        assert tuple(ref.media_type for ref in result.artifacts) == (
            XLSX_MEDIA_TYPE,
            "text/csv",
        )
    finally:
        await agent.close()


async def test_delivery_failure_after_xlsx_commit_retains_verified_artifact(
    tmp_path: Path,
) -> None:
    downloads = tmp_path / "downloads-xlsx-failure"
    downloads.mkdir()
    agent, provider, source_id, resource_id, _database = await _sqlite_export_agent(
        tmp_path,
        name="xlsx-delivery-failure",
        downloads=downloads,
        rows=(("retained", 1),),
    )
    provider.replace_script(
        (
            _tools(
                ToolCall(
                    id="export",
                    name=DATA_EXPORT_TABULAR_TOOL_NAME,
                    arguments={
                        "source_id": source_id,
                        "resource_ids": (resource_id,),
                        "sql": "SELECT label, number FROM records",
                        "format": "xlsx",
                        "filename": "retained.xlsx",
                    },
                )
            ),
            _tools(
                ToolCall(
                    id="delivery",
                    name="artifact_save_local",
                    arguments={
                        "artifact_id": "artifact-00000000000000000000000000000001",
                        "mode": "create_new",
                        "destination_id": "default",
                    },
                )
            ),
            _stop("delivery failed"),
        )
    )
    downloads.rmdir()
    try:
        result = await agent.run("Export this XLSX workbook.")
        assert len(result.artifacts) == 1
        assert result.artifact_deliveries == ()
        transcript = await agent.transcript(result.run_id)
        delivery = _result_for_call(transcript, "delivery")
        assert _error_code(delivery) == "artifact_downloads_unavailable"
        error = delivery.output["error"]
        assert isinstance(error, Mapping)
        details = error["details"]
        assert isinstance(details, Mapping)
        assert details["artifact_retained"] is True
        payload = await agent.read_artifact(result.artifacts[0].artifact_id)
        verify_exact_xlsx(payload.content)
    finally:
        await agent.close()


async def test_xlsx_cancellation_emits_no_artifact_bytes_reference_or_delivery(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    downloads = tmp_path / "downloads-xlsx-cancel"
    downloads.mkdir()
    agent, provider, source_id, resource_id, _database = await _sqlite_export_agent(
        tmp_path,
        name="xlsx-cancel",
        downloads=downloads,
    )
    started = threading.Event()
    release = threading.Event()

    def blocked(*args: Any, **kwargs: Any):
        del kwargs
        started.set()
        assert release.wait(2)
        return b"unused", ("label",), 0, f"schema_version:{args[5]}"

    monkeypatch.setattr(sqlite_query_module, "_run_exact_tabular_sync", blocked)
    provider.replace_script(
        (
            _tools(
                ToolCall(
                    id="cancelled-export",
                    name=DATA_EXPORT_TABULAR_TOOL_NAME,
                    arguments={
                        "source_id": source_id,
                        "resource_ids": (resource_id,),
                        "sql": "SELECT label FROM records",
                        "format": "xlsx",
                    },
                )
            ),
        )
    )
    try:
        running = asyncio.create_task(agent.run("Export this XLSX workbook."))
        assert await asyncio.to_thread(started.wait, 2)
        running.cancel()
        await asyncio.sleep(0)
        release.set()
        with pytest.raises(asyncio.CancelledError):
            await running
        assert not tuple(downloads.iterdir())
    finally:
        release.set()
        await agent.close()
