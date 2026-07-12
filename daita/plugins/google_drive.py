"""
Google Drive plugin for Daita Agents.

Read, search, and organize files across Google Drive with automatic format detection.
"""

import asyncio
import csv
import functools
import importlib
import io
import json
import logging
import mimetypes
import os
from pathlib import Path
from typing import (
    Any,
    Dict,
    List,
    Optional,
    Protocol,
    Sequence,
    TYPE_CHECKING,
    TypeGuard,
    TypedDict,
    TypeVar,
)

from .base import ConnectorPlugin, PluginContext
from .google_drive_extensions import (
    GOOGLE_DRIVE_MANIFEST,
    GoogleDriveExecutor,
    google_drive_capabilities,
    google_drive_evidence_schemas,
    google_drive_operation_definitions,
    google_drive_tool_views,
)
from ..core.exceptions import (
    PluginError,
    NotFoundError,
    PermissionError as DaitaPermissionError,
    AuthenticationError,
    TransientError,
)

if TYPE_CHECKING:
    from googleapiclient.http import HttpRequest

logger = logging.getLogger(__name__)

# Google Drive MIME types for native formats
_GDOC = "application/vnd.google-apps.document"
_GSHEET = "application/vnd.google-apps.spreadsheet"
_GSLIDES = "application/vnd.google-apps.presentation"
_GFOLDER = "application/vnd.google-apps.folder"

_FILE_TYPE_FILTERS = {
    "document": f"mimeType = '{_GDOC}'",
    "spreadsheet": f"mimeType = '{_GSHEET}'",
    "presentation": f"mimeType = '{_GSLIDES}'",
    "pdf": "mimeType = 'application/pdf'",
    "csv": "mimeType = 'text/csv'",
    "docx": "mimeType = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'",
    "xlsx": "mimeType = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
    "image": "mimeType contains 'image/'",
    "folder": f"mimeType = '{_GFOLDER}'",
}

_DEFAULT_SCOPES = ["https://www.googleapis.com/auth/drive.readonly"]
_WRITE_SCOPES = ["https://www.googleapis.com/auth/drive"]
_FILE_FIELDS = "id, name, mimeType, size, modifiedTime, parents, owners(displayName, emailAddress), webViewLink"
_LIST_FIELDS = f"files({_FILE_FIELDS}), nextPageToken"

# Export formats for Google-native files when downloading
_EXPORT_MAP = {
    _GDOC: (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        ".docx",
    ),
    _GSHEET: (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xlsx",
    ),
    _GSLIDES: (
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
        ".pptx",
    ),
}


class _DriveUser(TypedDict, total=False):
    displayName: str
    emailAddress: str


class _DriveFile(TypedDict, total=False):
    id: str
    name: str
    mimeType: str
    size: str
    modifiedTime: str
    createdTime: str
    parents: List[str]
    owners: List[_DriveUser]
    webViewLink: str
    description: str
    starred: bool
    trashed: bool
    lastModifyingUser: _DriveUser


class _DriveFileList(TypedDict, total=False):
    files: List[_DriveFile]
    nextPageToken: str


_ResponseT = TypeVar("_ResponseT", covariant=True)


class _DriveRequest(Protocol[_ResponseT]):
    def execute(self) -> _ResponseT: ...


class _DriveFilesResource(Protocol):
    def list(
        self,
        *,
        q: str,
        pageSize: int,
        fields: str,
        pageToken: Optional[str] = None,
    ) -> _DriveRequest[_DriveFileList]: ...

    def get(self, *, fileId: str, fields: str) -> _DriveRequest[_DriveFile]: ...

    def export(self, *, fileId: str, mimeType: str) -> _DriveRequest[bytes]: ...

    def get_media(self, *, fileId: str) -> "HttpRequest": ...

    def create(
        self,
        *,
        body: _DriveFile,
        media_body: object,
        fields: str,
    ) -> _DriveRequest[_DriveFile]: ...

    def update(
        self,
        *,
        fileId: str,
        body: _DriveFile,
        addParents: str,
        removeParents: str,
        fields: str,
    ) -> _DriveRequest[_DriveFile]: ...

    def copy(
        self, *, fileId: str, body: _DriveFile, fields: str
    ) -> _DriveRequest[_DriveFile]: ...

    def delete(self, *, fileId: str) -> _DriveRequest[object]: ...


class _DriveService(Protocol):
    def files(self) -> _DriveFilesResource: ...


def _is_drive_service(value: object) -> TypeGuard[_DriveService]:
    """Verify the one dynamic method exposed by Google's generated service."""
    return callable(getattr(value, "files", None))


class GoogleDrivePlugin(ConnectorPlugin):
    """
    Google Drive plugin for agents.

    Supports reading, searching, and organizing files across Google Drive.
    Handles Google-native formats (Docs, Sheets, Slides) and uploaded files
    (CSV, XLSX, DOCX, PDF, text) with automatic format detection.

    Auth:
        - Service account: credentials_path pointing to service account JSON
        - OAuth: credentials_path pointing to client secrets JSON (cached to token_path)
        - Pre-built: pass credentials object directly
        - ADC: omit all credential args (uses gcloud / GOOGLE_APPLICATION_CREDENTIALS)

    Example:
        async with google_drive(credentials_path="service_account.json") as drive:
            results = await drive.search("Q4 report")
            content = await drive.read(results[0]["id"])
    """

    _MAX_CHARS = 50_000
    _MAX_ROWS = 200
    manifest = GOOGLE_DRIVE_MANIFEST

    def __init__(
        self,
        credentials_path: Optional[str] = None,
        credentials: Optional[Any] = None,
        token_path: str = "~/.daita/gdrive_token.json",
        scopes: Optional[List[str]] = None,
        read_only: bool = True,
        **kwargs,
    ):
        """
        Initialize Google Drive plugin.

        Args:
            credentials_path: Path to service account JSON or OAuth client secrets JSON
            credentials: Pre-built google.auth Credentials object
            token_path: Path to cache OAuth tokens (for installed app flow)
            scopes: OAuth scopes. Defaults to drive.readonly (or drive if read_only=False)
            read_only: If True (default), only exposes read tools. Set False for upload/organize.
        """
        self.credentials_path = credentials_path
        self._credentials = credentials
        self.token_path = os.path.expanduser(token_path)
        self.read_only = read_only

        if scopes is not None:
            self.scopes = scopes
        else:
            self.scopes = _DEFAULT_SCOPES if read_only else _WRITE_SCOPES

        self._service: Optional[object] = None
        self._executor = GoogleDriveExecutor(self)
        self._connect_lock = asyncio.Lock()

        logger.debug("Google Drive plugin configured")

    @property
    def is_connected(self) -> bool:
        return self._service is not None

    @property
    def service(self) -> _DriveService:
        """Return the connected Drive service."""
        service = self._service
        if service is None:
            raise PluginError(
                "Google Drive is not connected", plugin_name="GoogleDrive"
            )
        if not _is_drive_service(service):
            raise PluginError(
                "Google Drive client has an invalid service interface",
                plugin_name="GoogleDrive",
            )
        return service

    async def setup(self, context: PluginContext) -> None:
        """Set up the Google Drive connector for a runtime."""
        await self.connect()

    async def teardown(self) -> None:
        """Disconnect the Google Drive connector from a runtime."""
        await self.disconnect()

    # ---------------------------------------------------------------------------
    # Runtime extension declarations
    # ---------------------------------------------------------------------------

    def declare_capabilities(self):
        return google_drive_capabilities(self.read_only)

    def get_executors(self):
        return (self._executor,)

    def declare_evidence_schemas(self):
        return google_drive_evidence_schemas()

    def get_tool_views(self):
        return google_drive_tool_views(self.read_only)

    def _definition_for_capability(self, capability_id: str) -> dict:
        for definition in google_drive_operation_definitions(self.read_only):
            if definition["capability_id"] == capability_id:
                return definition
        raise KeyError(capability_id)

    def _definition_for_tool(self, tool_name: str) -> dict:
        for definition in google_drive_operation_definitions(self.read_only):
            if definition["tool_name"] == tool_name:
                return definition
        raise KeyError(tool_name)

    # ---------------------------------------------------------------------------
    # Error mapping
    # ---------------------------------------------------------------------------

    def _map_drive_error(self, error: Exception, operation: str) -> Exception:
        """Map Google API errors to framework exceptions."""
        if isinstance(error, ImportError):
            return ImportError(
                "google-api-python-client is required for GoogleDrivePlugin. "
                "Install with: pip install 'daita-agents[google-drive]'"
            )
        try:
            from googleapiclient.errors import HttpError

            if isinstance(error, HttpError):
                status = error.resp.status
                if status == 404:
                    return NotFoundError(
                        f"Google Drive resource not found during {operation}",
                        resource_type="drive_file",
                    )
                elif status == 403:
                    return DaitaPermissionError(
                        f"Permission denied during Google Drive {operation}",
                        resource="drive",
                        action=operation,
                    )
                elif status == 401:
                    self._service = None  # Force reconnect on next call
                    return AuthenticationError(
                        f"Google Drive authentication failed during {operation}",
                        provider="Google Drive",
                    )
                elif status in (429, 500, 503):
                    return TransientError(
                        f"Google Drive {operation} transient error: {error}"
                    )
                else:
                    return PluginError(
                        f"Google Drive {operation} failed (HTTP {status}): {error}",
                        plugin_name="GoogleDrive",
                    )
        except ImportError:
            pass
        return PluginError(
            f"Google Drive {operation} failed: {error}", plugin_name="GoogleDrive"
        )

    # ---------------------------------------------------------------------------
    # Connection management
    # ---------------------------------------------------------------------------

    async def connect(self):
        """Initialize Google Drive service."""
        async with self._connect_lock:
            if self._service is not None:
                return

            try:
                from googleapiclient.discovery import build

                creds = await self._resolve_credentials()
                service = build("drive", "v3", credentials=creds, cache_discovery=False)

                # Verify connectivity with a minimal API call
                loop = asyncio.get_running_loop()
                await loop.run_in_executor(
                    None,
                    lambda: service.files()
                    .list(pageSize=1, fields="files(id)")
                    .execute(),
                )
                self._service = service
                logger.info("Connected to Google Drive")

            except ImportError as e:
                raise ImportError(
                    "google-api-python-client is required for GoogleDrivePlugin. "
                    "Install with: pip install 'daita-agents[google-drive]'"
                ) from e
            except (
                PluginError,
                NotFoundError,
                DaitaPermissionError,
                AuthenticationError,
                TransientError,
            ):
                raise
            except Exception as e:
                raise self._map_drive_error(e, "connect") from e

    async def _resolve_credentials(self) -> Any:
        """Resolve credentials from available sources (priority: object > file > ADC)."""
        if self._credentials is not None:
            return self._credentials

        if self.credentials_path:
            with open(self.credentials_path) as f:
                info = json.load(f)
            if info.get("type") == "service_account":
                from google.oauth2.service_account import Credentials

                return Credentials.from_service_account_file(
                    self.credentials_path, scopes=self.scopes
                )
            else:
                return await self._oauth_flow()

        # Application Default Credentials
        from google.auth import default as gauth_default
        from google.auth.exceptions import DefaultCredentialsError

        try:
            creds, _ = gauth_default(scopes=self.scopes)
            return creds
        except DefaultCredentialsError as error:
            raise AuthenticationError(
                "Google Drive credentials were not found. Configure Application "
                "Default Credentials, credentials_path, or credentials.",
                provider="Google Drive",
            ) from error

    async def _oauth_flow(self) -> Any:
        """Run OAuth installed app flow or load cached token."""
        from google.oauth2.credentials import Credentials as OAuthCredentials
        from google.auth.transport.requests import Request

        flow_module = importlib.import_module("google_auth_oauthlib.flow")
        flow_class: object = getattr(flow_module, "InstalledAppFlow", None)
        from_client_secrets_file = getattr(flow_class, "from_client_secrets_file", None)
        if not callable(from_client_secrets_file):
            raise ImportError(
                "google-auth-oauthlib is required. Install with: "
                "pip install 'daita-agents[google-drive]'"
            )

        creds = None
        if os.path.exists(self.token_path):
            creds = OAuthCredentials.from_authorized_user_file(
                self.token_path, self.scopes
            )

        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                credentials_path = self.credentials_path
                if not credentials_path:
                    raise AuthenticationError(
                        "Google Drive OAuth requires credentials_path",
                        provider="Google Drive",
                    )
                flow = from_client_secrets_file(credentials_path, self.scopes)
                run_local_server = getattr(flow, "run_local_server", None)
                if not callable(run_local_server):
                    raise PluginError(
                        "Google OAuth flow has an invalid interface",
                        plugin_name="GoogleDrive",
                    )
                creds = run_local_server(port=0)

            os.makedirs(os.path.dirname(self.token_path), exist_ok=True)
            to_json = getattr(creds, "to_json", None)
            if not callable(to_json):
                raise PluginError(
                    "Google OAuth credentials have an invalid interface",
                    plugin_name="GoogleDrive",
                )
            token_json = to_json()
            if not isinstance(token_json, str):
                raise PluginError(
                    "Google OAuth credentials returned an invalid token payload",
                    plugin_name="GoogleDrive",
                )
            with open(self.token_path, "w") as f:
                f.write(token_json)

        return creds

    async def disconnect(self):
        """Close Google Drive connection."""
        if self._service:
            self._service = None
            logger.info("Disconnected from Google Drive")

    # ---------------------------------------------------------------------------
    # Sync helpers (called via run_in_executor — no inline lambdas)
    # ---------------------------------------------------------------------------

    def _sync_files_list(
        self, query: str, page_size: int, fields: str, page_token: Optional[str]
    ) -> _DriveFileList:
        return (
            self.service.files()
            .list(
                q=query,
                pageSize=page_size,
                fields=fields,
                pageToken=page_token,
            )
            .execute()
        )

    def _sync_files_get(self, file_id: str, fields: str) -> _DriveFile:
        return self.service.files().get(fileId=file_id, fields=fields).execute()

    def _sync_files_export(self, file_id: str, mime_type: str) -> bytes:
        return self.service.files().export(fileId=file_id, mimeType=mime_type).execute()

    def _sync_files_download(self, file_id: str) -> bytes:
        from googleapiclient.http import MediaIoBaseDownload

        request = self.service.files().get_media(fileId=file_id)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        return buf.getvalue()

    def _sync_files_create(
        self, metadata: _DriveFile, media_body: object
    ) -> _DriveFile:
        return (
            self.service.files()
            .create(body=metadata, media_body=media_body, fields=_FILE_FIELDS)
            .execute()
        )

    def _sync_files_update(
        self,
        file_id: str,
        metadata: _DriveFile,
        add_parents: str,
        remove_parents: str,
    ) -> _DriveFile:
        return (
            self.service.files()
            .update(
                fileId=file_id,
                body=metadata,
                addParents=add_parents,
                removeParents=remove_parents,
                fields=_FILE_FIELDS,
            )
            .execute()
        )

    def _sync_files_copy(self, file_id: str, metadata: _DriveFile) -> _DriveFile:
        return (
            self.service.files()
            .copy(fileId=file_id, body=metadata, fields=_FILE_FIELDS)
            .execute()
        )

    def _sync_files_delete(self, file_id: str) -> None:
        self.service.files().delete(fileId=file_id).execute()

    # ---------------------------------------------------------------------------
    # Core operations
    # ---------------------------------------------------------------------------

    async def search(
        self,
        query: str,
        file_type: Optional[str] = None,
        folder_id: Optional[str] = None,
        modified_after: Optional[str] = None,
        max_results: int = 20,
    ) -> List[Dict[str, Any]]:
        """Search Google Drive files by name, type, or date."""
        if self._service is None:
            await self.connect()

        parts = [f"name contains '{query}' and trashed = false"]
        if file_type and file_type in _FILE_TYPE_FILTERS:
            parts.append(_FILE_TYPE_FILTERS[file_type])
        if folder_id:
            parts.append(f"'{folder_id}' in parents")
        if modified_after:
            parts.append(f"modifiedTime > '{modified_after}T00:00:00'")

        drive_query = " and ".join(parts)
        page_size = min(max_results, 100)

        try:
            loop = asyncio.get_running_loop()
            response = await loop.run_in_executor(
                None,
                functools.partial(
                    self._sync_files_list, drive_query, page_size, _LIST_FIELDS, None
                ),
            )
            return self._format_file_list(response.get("files", []))
        except (
            PluginError,
            NotFoundError,
            DaitaPermissionError,
            AuthenticationError,
            TransientError,
        ):
            raise
        except Exception as e:
            raise self._map_drive_error(e, "search") from e

    async def list_folder(
        self,
        folder_id: str = "root",
        max_results: int = 50,
    ) -> List[Dict[str, Any]]:
        """List files in a Drive folder."""
        if self._service is None:
            await self.connect()

        query = f"'{folder_id}' in parents and trashed = false"
        try:
            loop = asyncio.get_running_loop()
            response = await loop.run_in_executor(
                None,
                functools.partial(
                    self._sync_files_list,
                    query,
                    min(max_results, 200),
                    _LIST_FIELDS,
                    None,
                ),
            )
            return self._format_file_list(response.get("files", []))
        except (
            PluginError,
            NotFoundError,
            DaitaPermissionError,
            AuthenticationError,
            TransientError,
        ):
            raise
        except Exception as e:
            raise self._map_drive_error(e, "list_folder") from e

    async def get_info(self, file_id: str) -> Dict[str, Any]:
        """Get detailed file metadata."""
        if self._service is None:
            await self.connect()

        fields = (
            "id, name, mimeType, size, modifiedTime, createdTime, parents, "
            "owners(displayName, emailAddress), webViewLink, description, "
            "starred, trashed, lastModifyingUser(displayName, emailAddress)"
        )
        try:
            loop = asyncio.get_running_loop()
            f = await loop.run_in_executor(
                None,
                functools.partial(self._sync_files_get, file_id, fields),
            )
            return {
                "id": f.get("id"),
                "name": f.get("name"),
                "type": self._friendly_type(f.get("mimeType", "")),
                "mime_type": f.get("mimeType"),
                "size": f.get("size"),
                "created": f.get("createdTime"),
                "modified": f.get("modifiedTime"),
                "owners": [o.get("emailAddress") for o in f.get("owners", [])],
                "last_modified_by": f.get("lastModifyingUser", {}).get("emailAddress"),
                "web_link": f.get("webViewLink"),
                "description": f.get("description"),
                "starred": f.get("starred"),
            }
        except (
            PluginError,
            NotFoundError,
            DaitaPermissionError,
            AuthenticationError,
            TransientError,
        ):
            raise
        except Exception as e:
            raise self._map_drive_error(e, "get_info") from e

    async def read(
        self, file_id: str, sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """Read file content with automatic format detection."""
        if self._service is None:
            await self.connect()

        try:
            loop = asyncio.get_running_loop()
            meta = await loop.run_in_executor(
                None,
                functools.partial(
                    self._sync_files_get, file_id, "id, name, mimeType, size"
                ),
            )
            return await self._extract_content(
                file_id, meta.get("mimeType", ""), meta.get("name", ""), sheet_name
            )
        except (
            PluginError,
            NotFoundError,
            DaitaPermissionError,
            AuthenticationError,
            TransientError,
        ):
            raise
        except Exception as e:
            raise self._map_drive_error(e, "read") from e

    async def download(self, file_id: str, local_path: str) -> str:
        """Download file to local path. Google-native formats are auto-exported."""
        if self._service is None:
            await self.connect()

        try:
            loop = asyncio.get_running_loop()
            meta = await loop.run_in_executor(
                None,
                functools.partial(self._sync_files_get, file_id, "id, name, mimeType"),
            )
            mime_type = meta.get("mimeType", "")
            dest = os.path.expanduser(local_path)

            if mime_type in _EXPORT_MAP:
                export_mime, ext = _EXPORT_MAP[mime_type]
                if not dest.endswith(ext):
                    dest = dest + ext
                content = await loop.run_in_executor(
                    None,
                    functools.partial(self._sync_files_export, file_id, export_mime),
                )
            else:
                content = await loop.run_in_executor(
                    None,
                    functools.partial(self._sync_files_download, file_id),
                )

            os.makedirs(os.path.dirname(os.path.abspath(dest)), exist_ok=True)
            with open(dest, "wb") as f:
                f.write(content)

            return dest
        except (
            PluginError,
            NotFoundError,
            DaitaPermissionError,
            AuthenticationError,
            TransientError,
        ):
            raise
        except Exception as e:
            raise self._map_drive_error(e, "download") from e

    async def upload(
        self,
        local_path: str,
        folder_id: Optional[str] = None,
        name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Upload a local file to Google Drive."""
        if self._service is None:
            await self.connect()

        local_path = os.path.expanduser(local_path)
        if not os.path.exists(local_path):
            raise NotFoundError(
                f"Local file not found: {local_path}", resource_type="local_file"
            )

        try:
            from googleapiclient.http import MediaFileUpload

            file_name = name or Path(local_path).name
            metadata: _DriveFile = {"name": file_name}
            if folder_id:
                metadata["parents"] = [folder_id]

            mime_type, _ = mimetypes.guess_type(local_path)
            media = MediaFileUpload(
                local_path, mimetype=mime_type or "application/octet-stream"
            )

            loop = asyncio.get_running_loop()
            result = await loop.run_in_executor(
                None,
                functools.partial(self._sync_files_create, metadata, media),
            )
            return self._format_file(result)
        except (
            PluginError,
            NotFoundError,
            DaitaPermissionError,
            AuthenticationError,
            TransientError,
        ):
            raise
        except Exception as e:
            raise self._map_drive_error(e, "upload") from e

    async def organize(
        self,
        file_id: str,
        action: str,
        dest_folder_id: Optional[str] = None,
        new_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Move, rename, or copy a file in Google Drive."""
        if self._service is None:
            await self.connect()

        if action not in ("move", "rename", "copy"):
            raise PluginError(
                f"Invalid action '{action}'. Must be one of: move, rename, copy",
                plugin_name="GoogleDrive",
            )

        try:
            loop = asyncio.get_running_loop()

            if action == "copy":
                metadata: _DriveFile = {}
                if new_name:
                    metadata["name"] = new_name
                if dest_folder_id:
                    metadata["parents"] = [dest_folder_id]
                result = await loop.run_in_executor(
                    None,
                    functools.partial(self._sync_files_copy, file_id, metadata),
                )
                return {**self._format_file(result), "action": "copied"}

            elif action == "rename":
                if not new_name:
                    raise PluginError(
                        "rename requires new_name", plugin_name="GoogleDrive"
                    )
                result = await loop.run_in_executor(
                    None,
                    functools.partial(
                        self._sync_files_update, file_id, {"name": new_name}, "", ""
                    ),
                )
                return {**self._format_file(result), "action": "renamed"}

            else:  # move
                if not dest_folder_id:
                    raise PluginError(
                        "move requires dest_folder_id", plugin_name="GoogleDrive"
                    )
                meta = await loop.run_in_executor(
                    None,
                    functools.partial(self._sync_files_get, file_id, "parents"),
                )
                current_parents = ",".join(meta.get("parents", []))
                result = await loop.run_in_executor(
                    None,
                    functools.partial(
                        self._sync_files_update,
                        file_id,
                        {},
                        dest_folder_id,
                        current_parents,
                    ),
                )
                return {**self._format_file(result), "action": "moved"}

        except (
            PluginError,
            NotFoundError,
            DaitaPermissionError,
            AuthenticationError,
            TransientError,
        ):
            raise
        except Exception as e:
            raise self._map_drive_error(e, "organize") from e

    # ---------------------------------------------------------------------------
    # Content extraction (the "wow" — pass any file ID, get structured content)
    # ---------------------------------------------------------------------------

    async def _extract_content(
        self,
        file_id: str,
        mime_type: str,
        name: str,
        sheet_name: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Extract readable content from a Drive file based on MIME type."""
        loop = asyncio.get_running_loop()

        # Google Docs → plain text
        if mime_type == _GDOC:
            raw = await loop.run_in_executor(
                None,
                functools.partial(self._sync_files_export, file_id, "text/plain"),
            )
            return self._wrap_text(raw.decode("utf-8"), name, "google_doc")

        # Google Sheets → CSV rows
        if mime_type == _GSHEET:
            raw = await loop.run_in_executor(
                None,
                functools.partial(self._sync_files_export, file_id, "text/csv"),
            )
            rows = list(csv.DictReader(io.StringIO(raw.decode("utf-8"))))
            return self._wrap_rows(rows, name, "google_sheet")

        # Google Slides → plain text (slide content extracted)
        if mime_type == _GSLIDES:
            raw = await loop.run_in_executor(
                None,
                functools.partial(self._sync_files_export, file_id, "text/plain"),
            )
            return self._wrap_text(raw.decode("utf-8"), name, "google_slides")

        # Folder — no content
        if mime_type == _GFOLDER:
            return {
                "file_id": file_id,
                "name": name,
                "type": "folder",
                "binary": True,
                "note": "This is a folder. Use gdrive_list to see its contents.",
            }

        # Uploaded files — download bytes and parse by MIME type / extension
        raw = await loop.run_in_executor(
            None,
            functools.partial(self._sync_files_download, file_id),
        )

        low_name = name.lower()

        if mime_type == "text/csv" or low_name.endswith(".csv"):
            rows = list(csv.DictReader(io.StringIO(raw.decode("utf-8"))))
            return self._wrap_rows(rows, name, "csv")

        if (
            mime_type
            == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            or low_name.endswith(".xlsx")
        ):
            return self._parse_xlsx(raw, name, sheet_name)

        if (
            mime_type
            == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            or low_name.endswith(".docx")
        ):
            return self._parse_docx(raw, name)

        if mime_type == "application/pdf" or low_name.endswith(".pdf"):
            return self._parse_pdf(raw, name)

        if mime_type == "application/json" or low_name.endswith(".json"):
            try:
                data = json.loads(raw.decode("utf-8"))
                serialized = json.dumps(data, indent=2)
                truncated = len(serialized) > self._MAX_CHARS
                return {
                    "name": name,
                    "format": "json",
                    "content": serialized[: self._MAX_CHARS],
                    "truncated": truncated,
                    **({"total_chars": len(serialized)} if truncated else {}),
                }
            except (UnicodeDecodeError, json.JSONDecodeError):
                pass

        if mime_type.startswith("text/") or low_name.endswith(
            (
                ".txt",
                ".md",
                ".xml",
                ".yaml",
                ".yml",
                ".log",
                ".py",
                ".js",
                ".ts",
                ".sql",
            )
        ):
            try:
                return self._wrap_text(raw.decode("utf-8"), name, "text")
            except UnicodeDecodeError:
                pass

        # Unrecognized binary
        return {
            "file_id": file_id,
            "name": name,
            "mime_type": mime_type,
            "size": len(raw),
            "binary": True,
            "note": "Binary file — content not extractable. Use gdrive_download to save locally.",
        }

    def _parse_xlsx(
        self, content: bytes, name: str, sheet_name: Optional[str] = None
    ) -> Dict[str, Any]:
        """Parse XLSX bytes using openpyxl (from [data] extra)."""
        try:
            import openpyxl
        except ImportError:
            return {
                "name": name,
                "binary": True,
                "note": "Install daita-agents[data] for XLSX content extraction.",
            }

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

        if ws is None:
            return self._wrap_rows([], name, "xlsx")

        rows = []
        headers: Optional[List[str]] = None
        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = [str(c) if c is not None else "" for c in row]
                continue
            rows.append(
                dict(zip(headers, [str(v) if v is not None else "" for v in row]))
            )

        return self._wrap_rows(rows, name, "xlsx")

    def _parse_docx(self, content: bytes, name: str) -> Dict[str, Any]:
        """Parse DOCX bytes using python-docx."""
        try:
            import docx
        except ImportError:
            raise ImportError(
                "python-docx is required. Install with: pip install 'daita-agents[google-drive]'"
            )
        doc = docx.Document(io.BytesIO(content))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return self._wrap_text(text, name, "docx")

    def _parse_pdf(self, content: bytes, name: str) -> Dict[str, Any]:
        """Extract text from PDF using pypdf."""
        try:
            import pypdf
        except ImportError:
            raise ImportError(
                "pypdf is required. Install with: pip install 'daita-agents[google-drive]'"
            )
        reader = pypdf.PdfReader(io.BytesIO(content))
        pages = [page.extract_text() or "" for page in reader.pages]
        full_text = "\n\n".join(p for p in pages if p.strip())
        return self._wrap_text(full_text, name, "pdf")

    def _wrap_text(self, text: str, name: str, fmt: str) -> Dict[str, Any]:
        truncated = len(text) > self._MAX_CHARS
        return {
            "name": name,
            "format": fmt,
            "content": text[: self._MAX_CHARS],
            "truncated": truncated,
            **({"total_chars": len(text)} if truncated else {}),
        }

    def _wrap_rows(self, rows: List[Dict], name: str, fmt: str) -> Dict[str, Any]:
        total = len(rows)
        truncated = total > self._MAX_ROWS
        return {
            "name": name,
            "format": fmt,
            "rows": rows[: self._MAX_ROWS],
            "total_rows": total,
            "truncated": truncated,
        }

    # ---------------------------------------------------------------------------
    # Formatting helpers
    # ---------------------------------------------------------------------------

    def _format_file(self, f: _DriveFile) -> Dict[str, Any]:
        """Full metadata for single-file results (upload, organize)."""
        return {
            "id": f.get("id"),
            "name": f.get("name"),
            "type": self._friendly_type(f.get("mimeType", "")),
            "mime_type": f.get("mimeType"),
            "size": f.get("size"),
            "modified": f.get("modifiedTime"),
            "owners": [o.get("emailAddress") for o in f.get("owners", [])],
            "web_link": f.get("webViewLink"),
        }

    def _format_file_slim(self, f: _DriveFile) -> Dict[str, Any]:
        """Minimal fields for list/search results — keeps token cost low."""
        return {
            "id": f.get("id"),
            "name": f.get("name"),
            "type": self._friendly_type(f.get("mimeType", "")),
            "modified": f.get("modifiedTime"),
        }

    def _format_file_list(self, files: Sequence[_DriveFile]) -> List[Dict[str, Any]]:
        return [self._format_file_slim(f) for f in files]

    def _friendly_type(self, mime_type: str) -> str:
        return {
            _GDOC: "Google Doc",
            _GSHEET: "Google Sheet",
            _GSLIDES: "Google Slides",
            _GFOLDER: "Folder",
            "application/pdf": "PDF",
            "text/csv": "CSV",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document": "DOCX",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": "XLSX",
            "image/png": "PNG",
            "image/jpeg": "JPEG",
        }.get(mime_type, mime_type)

    # ---------------------------------------------------------------------------
    # Tool definitions
    # ---------------------------------------------------------------------------

    # ---------------------------------------------------------------------------
    # Tool handlers
    # ---------------------------------------------------------------------------

    async def _tool_search(self, args: Dict[str, Any]) -> Dict[str, Any]:
        results = await self.search(
            query=args["query"],
            file_type=args.get("file_type"),
            folder_id=args.get("folder_id"),
            modified_after=args.get("modified_after"),
            max_results=args.get("max_results", 20),
        )
        return {"files": results, "count": len(results)}

    async def _tool_read(self, args: Dict[str, Any]) -> Dict[str, Any]:
        return await self.read(
            file_id=args["file_id"], sheet_name=args.get("sheet_name")
        )

    async def _tool_list(self, args: Dict[str, Any]) -> Dict[str, Any]:
        files = await self.list_folder(
            folder_id=args.get("folder_id", "root"),
            max_results=args.get("max_results", 50),
        )
        return {
            "files": files,
            "count": len(files),
            "folder_id": args.get("folder_id", "root"),
        }

    async def _tool_info(self, args: Dict[str, Any]) -> Dict[str, Any]:
        return await self.get_info(file_id=args["file_id"])

    async def _tool_download(self, args: Dict[str, Any]) -> Dict[str, Any]:
        saved_path = await self.download(
            file_id=args["file_id"],
            local_path=args["local_path"],
        )
        return {"saved_to": saved_path, "file_id": args["file_id"]}

    async def _tool_upload(self, args: Dict[str, Any]) -> Dict[str, Any]:
        result = await self.upload(
            local_path=args["local_path"],
            folder_id=args.get("folder_id"),
            name=args.get("name"),
        )
        return {**result, "uploaded": True}

    async def _tool_organize(self, args: Dict[str, Any]) -> Dict[str, Any]:
        return await self.organize(
            file_id=args["file_id"],
            action=args["action"],
            dest_folder_id=args.get("dest_folder_id"),
            new_name=args.get("new_name"),
        )

    # Context manager support
    async def __aenter__(self):
        await self.connect()
        return self

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        await self.disconnect()


def google_drive(**kwargs) -> GoogleDrivePlugin:
    """Create Google Drive plugin."""
    return GoogleDrivePlugin(**kwargs)
